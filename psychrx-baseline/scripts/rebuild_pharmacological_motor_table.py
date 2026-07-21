"""Rebuild the pharmacological motor table from the audited workbook tabs."""

from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TABLE_DIR = ROOT / "knowledge_base" / "decision_support_engine" / "tables"
WORKBOOK = TABLE_DIR / "Tabela_Motor_Abas_Unificadas_audit_evilasio.xlsx"
OUTPUT_WORKBOOK = TABLE_DIR / "Tabela_Motor_Abas_Unificadas_audit_evilasio_CORRIGIDA.xlsx"
CSV_PATH = TABLE_DIR / "pharmacological_decision_matrix.csv"

HEADERS = [
    "drug_id",
    "drug_name",
    "drug_class",
    "therapeutic_targets",
    "fit_energy",
    "fit_anxiety",
    "fit_sleep",
    "fit_pain",
    "fit_weight_neutral",
    "fit_libido",
    "fit_mood_stabilization",
    "fit_psychosis",
    "fit_impulsivity",
    "cautions",
    "preferred_contexts",
    "source_tabs",
    "status",
]

ALIASES = {
    "Desvenlafaxina 50-100": "Desvenlafaxina",
    "Duloxetina 60-120": "Duloxetina",
    "Citalopram 20-40": "Citalopram",
    "Escitalopram 10-20": "Escitalopram",
    "Bupropiona 150-300": "Bupropiona",
    "Venlafaxina": "Venlafaxina",
    "Venlafaxina XR": "Venlafaxina XR",
    "Haloperidol (Haldol)": "Haloperidol",
    "Clorpromazina (Amplictil)": "Clorpromazina",
    "Levomepromazina (Neozine)": "Levomepromazina",
    "Tioridazina (neuliptil)": "Tioridazina",
    "fenobrabital": "Fenobarbital",
    "fenitoina": "Fenitoina",
    "cabamazepina": "Carbamazepina",
    "oxcabamazepina": "Oxcarbazepina",
    "Valproato de sódio": "Valproato",
    "Carbonato de lítio": "Litio",
    "lítio": "Litio",
    "valproato": "Valproato",
    "divalproato": "Divalproato",
    "depakote": "Divalproato",
    "torval": "Valproato",
    "carbamazepina": "Carbamazepina",
    "lamotrigina": "Lamotrigina",
    "lurasidona": "Lurasidona",
    "topiramato": "Topiramato",
    "naltrexona": "Naltrexona",
    "uninaltrex": "Naltrexona",
}

NON_DRUG_TERMS = {
    "anticonvulsivantes",
    "antiepilépticos",
    "antiepilepticos",
}

MANUAL = {
    "Sertralina": ("SSRI", "ansiedade|humor|toc", 2, 4, 1, 1, 3, 1, 1, 0, 1, "disfuncao sexual|gi|ativacao inicial", "ansiedade|toc|depressiva"),
    "Escitalopram": ("SSRI", "ansiedade|humor", 2, 4, 1, 1, 3, 1, 1, 0, 1, "qt|disfuncao sexual", "ansiedade|depressiva"),
    "Fluoxetina": ("SSRI", "humor|energia", 3, 3, 0, 1, 4, 1, 1, 0, 2, "ativacao|insonia|disfuncao sexual", "depressiva|apatica"),
    "Venlafaxina": ("SNRI", "energia|ansiedade|humor|dor", 3, 4, 1, 3, 3, 1, 1, 0, 2, "pressao arterial|retirada|ativacao", "depressiva|ansiosa|dor"),
    "Venlafaxina XR": ("SNRI", "energia|ansiedade|humor|dor", 4, 4, 1, 3, 3, 1, 1, 0, 2, "pressao arterial|retirada|ativacao", "depressiva|ansiosa|dor"),
    "Desvenlafaxina": ("SNRI", "energia|ansiedade|humor", 4, 3, 1, 2, 3, 1, 1, 0, 2, "pressao arterial|retirada", "depressiva|ansiosa"),
    "Duloxetina": ("SNRI", "dor|ansiedade|humor|energia", 4, 4, 1, 5, 3, 1, 1, 0, 1, "hepatico|renal|pressao arterial", "dor|ansiosa|somatica"),
    "Bupropiona": ("NDRI", "energia|humor|libido|tabagismo", 4, 1, 0, 1, 5, 5, 1, 0, 3, "convulsao|transtorno alimentar|insonia|ansiedade", "apatica|fadiga|libido|tabagismo"),
    "Bupropiona XL": ("NDRI", "energia|humor|libido|tabagismo", 5, 1, 0, 1, 5, 5, 1, 0, 3, "convulsao|transtorno alimentar|insonia|ansiedade", "apatica|fadiga|libido|tabagismo"),
    "Mirtazapina": ("NaSSA", "sono|apetite|humor", 1, 3, 5, 1, 0, 4, 1, 0, 1, "ganho de peso|sedacao|metabolico", "insonia|perda de apetite|ansiosa"),
    "Trazodona": ("SARI", "sono|humor", 1, 2, 5, 1, 2, 3, 1, 0, 1, "sedacao|hipotensao", "sono|insonia"),
    "Risperidona": ("Antipsicotico atipico", "psicose|mania|irritabilidade", 1, 1, 2, 0, 1, 0, 3, 5, 4, "prolactina|eps|metabolico", "psicotica|maniaca"),
    "Olanzapina": ("Antipsicotico atipico", "psicose|mania|sono|apetite", 1, 2, 4, 0, 0, 1, 4, 5, 3, "ganho de peso|metabolico|sedacao", "psicotica|maniaca|agitacao"),
    "Quetiapina": ("Antipsicotico atipico", "sono|psicose|bipolar|ansiedade", 1, 3, 5, 0, 1, 2, 4, 4, 2, "sedacao|metabolico|hipotensao", "sono|bipolar|psicotica"),
    "Aripiprazol": ("Antipsicotico atipico", "psicose|mania|potencializacao|prolactina", 3, 1, 0, 0, 4, 3, 4, 4, 4, "acatisia|ativacao|insonia", "psicotica|maniaca|prolactina"),
    "Lurasidona": ("Antipsicotico atipico", "bipolar|psicose|humor|ansiedade", 2, 2, 1, 0, 4, 3, 3, 4, 2, "acatisia|tomar com alimento", "bipolar|psicotica|ansiosa"),
    "Litio": ("Estabilizador de humor", "mania|suicidio|estabilidade", 1, 1, 1, 0, 2, 2, 5, 1, 4, "renal|tireoide|nivel serico|interacoes", "maniaca|recorrencia|suicida"),
    "Lamotrigina": ("Estabilizador de humor", "depressao bipolar|estabilidade", 3, 2, 1, 0, 5, 4, 4, 0, 2, "rash|titulacao lenta", "bipolar|depressao bipolar"),
    "Valproato": ("Estabilizador de humor", "mania|impulsividade|estabilidade", 1, 1, 2, 0, 0, 1, 5, 1, 4, "gestacao|hepatico|peso|plaquetas", "maniaca|impulsividade"),
    "Carbamazepina": ("Estabilizador de humor", "mania|impulsividade|estabilidade", 1, 1, 1, 0, 2, 1, 4, 1, 3, "interacoes|hematologico|hepatico", "maniaca|impulsividade"),
}


def slug(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def clean_name(raw: str) -> str | None:
    value = " ".join(str(raw).strip().split())
    if not value or value.lower() in {"medicamento", "antidepressivos", "cor", "uso"}:
        return None
    if "registrar fármaco" in value.lower() or "registrar farmaco" in value.lower():
        return None
    if value.lower() in NON_DRUG_TERMS:
        return None
    return ALIASES.get(value, value)


def add_drug(rows: dict[str, dict[str, object]], name: str, source: str) -> None:
    clean = clean_name(name)
    if not clean:
        return
    key = slug(clean)
    row = rows.setdefault(
        key,
        {
            "drug_id": key,
            "drug_name": clean,
            "drug_class": "",
            "therapeutic_targets": "",
            "fit_energy": 0,
            "fit_anxiety": 0,
            "fit_sleep": 0,
            "fit_pain": 0,
            "fit_weight_neutral": 0,
            "fit_libido": 0,
            "fit_mood_stabilization": 0,
            "fit_psychosis": 0,
            "fit_impulsivity": 0,
            "cautions": "",
            "preferred_contexts": "",
            "source_tabs": "",
            "status": "draft_from_source_audit",
        },
    )
    sources = set(filter(None, str(row["source_tabs"]).split("|")))
    sources.add(source)
    row["source_tabs"] = "|".join(sorted(sources))


def apply_profile(row: dict[str, object]) -> None:
    name = str(row["drug_name"])
    profile = MANUAL.get(name)
    if not profile:
        # Conservative source-derived defaults by broad class.
        if name in {"Amitriptilina", "Nortriptilina", "Clomipramina", "Imipramina", "Desipramina", "Doxepina", "Dosulepina", "Maprotilina"}:
            profile = ("Triciclico", "humor|dor|sono", 2, 2, 3, 2, 1, 1, 1, 0, 1, "anticolinergico|sedacao|cardiaco|toxicidade", "depressiva|dor|sono")
        elif name in {"Citalopram", "Fluvoxamina", "Paroxetina", "Vortioxetina", "Vilazodona"}:
            profile = ("SSRI/serotonergico", "ansiedade|humor", 2, 3, 1, 1, 3, 1, 1, 0, 1, "disfuncao sexual|gi|qt conforme farmaco", "ansiedade|depressiva")
        elif name in {"Agomelatina"}:
            profile = ("Atipico antidepressivo", "sono|humor", 2, 2, 3, 1, 4, 3, 1, 0, 1, "hepatico|monitorizacao", "sono|depressiva")
        elif name in {"Mianserina"}:
            profile = ("Tetraciclico", "sono|humor", 1, 2, 4, 1, 1, 3, 1, 0, 1, "sedacao|peso", "sono|depressiva")
        elif name in {"Haloperidol", "Clorpromazina", "Levomepromazina", "Tioridazina"}:
            profile = ("Antipsicotico tipico", "psicose|agitacao|sedacao", 0, 1, 2, 0, 1, 0, 2, 5, 3, "eps|anticolinergico|prolactina|qt|sedacao", "psicotica|agitacao")
        elif name in {"Clozapina"}:
            profile = ("Antipsicotico atipico", "psicose|refratariedade", 0, 1, 4, 0, 0, 1, 3, 5, 2, "metabolico|sedacao|agranulocitose|monitorizacao", "psicotica|refrataria")
        elif name in {"Fenobarbital", "Fenitoina", "Oxcarbazepina", "Divalproato", "Topiramato"}:
            profile = ("Anticonvulsivante/estabilizador", "estabilidade|impulsividade", 1, 1, 1, 0, 2, 1, 3, 0, 3, "interacoes|sedacao|monitorizacao", "estabilidade|impulsividade")
        elif name in {"Naltrexona"}:
            profile = ("Dependencia quimica", "compulsao|substancias", 1, 1, 0, 0, 4, 3, 1, 0, 3, "hepatico|opioides|substancias", "dependencia|compulsao")
        else:
            profile = ("Classe pendente", "alvo pendente", 0, 0, 0, 0, 0, 0, 0, 0, 0, "revisao manual", "revisao")
    (
        row["drug_class"],
        row["therapeutic_targets"],
        row["fit_energy"],
        row["fit_anxiety"],
        row["fit_sleep"],
        row["fit_pain"],
        row["fit_weight_neutral"],
        row["fit_libido"],
        row["fit_mood_stabilization"],
        row["fit_psychosis"],
        row["fit_impulsivity"],
        row["cautions"],
        row["preferred_contexts"],
    ) = profile


def main() -> None:
    wb = load_workbook(WORKBOOK)
    rows: dict[str, dict[str, object]] = {}

    source_cols = {
        "Tabela 1 - Matriz Antidepressiv": (1, "Tabela 1"),
        "Tabela 2 - Antidepressivos": (10, "Tabela 2"),
        "Tabela 3 - Antipsicoticos Tipic": (9, "Tabela 3"),
        "Tabela 5 - Atipicos": (8, "Tabela 5"),
        "Tabela 7 - Atipicos Continuacao": (8, "Tabela 7"),
    }
    for sheet, (col, source) in source_cols.items():
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            value = ws.cell(r, col).value
            if isinstance(value, str):
                add_drug(rows, value, source)

    ws8 = wb["Tabela 8 - Estabilizadores"]
    for r in range(2, ws8.max_row + 1):
        value = ws8.cell(r, 7).value
        if not isinstance(value, str):
            continue
        for part in re.split(r"[;,]", value):
            item = part.strip()
            if item.lower().startswith("exce"):
                item = item.split(":", 1)[-1].strip()
            add_drug(rows, item, "Tabela 8")

    # Preserve formulation names already consumed by the app while keeping the
    # source-level base names extracted from the audited tabs.
    if "venlafaxina" in rows:
        add_drug(rows, "Venlafaxina XR", "Tabela 1|Tabela 2")
    if "bupropiona" in rows:
        add_drug(rows, "Bupropiona XL", "Tabela 1|Tabela 2")

    for row in rows.values():
        apply_profile(row)

    ordered = sorted(rows.values(), key=lambda row: (str(row["drug_class"]), str(row["drug_name"])))
    ws_name = "Aba 10 - Motor Decisao"
    if ws_name in wb.sheetnames:
        del wb[ws_name]
    ws = wb.create_sheet(ws_name)
    header_fill = PatternFill("solid", fgColor="0F8475")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r, row in enumerate(ordered, 2):
        for c, header in enumerate(HEADERS, 1):
            ws.cell(r, c, row[header])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(HEADERS[col - 1]) + 2, 12), 26)
    try:
        wb.save(WORKBOOK)
        saved_path = WORKBOOK
    except PermissionError:
        wb.save(OUTPUT_WORKBOOK)
        saved_path = OUTPUT_WORKBOOK

    with CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(ordered)

    print(f"Rebuilt {len(ordered)} motor rows")
    print(f"Workbook saved: {saved_path}")
    print(f"CSV saved: {CSV_PATH}")
    print("\\n".join(f"- {row['drug_name']} [{row['source_tabs']}]" for row in ordered))


if __name__ == "__main__":
    main()
